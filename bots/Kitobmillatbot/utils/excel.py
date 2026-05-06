import csv
import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def export_users_to_excel(users: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"

    headers = [
        "ID",
        "FIO",
        "Username",
        "Telefon",
        "Referallar",
        "Ball",
        "Javob",
        "Kim taklif qildi (ID)",
        "Telegram ID raqami",
        "qo'shilgan vaqti",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, user in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=user.telegram_id)
        ws.cell(row=row_idx, column=2, value=user.fio or "")
        ws.cell(row=row_idx, column=3, value=f"@{user.username}" if user.username else "")
        ws.cell(row=row_idx, column=4, value=user.mobile_number or "")
        ws.cell(row=row_idx, column=5, value=user.referrals_count)
        ws.cell(row=row_idx, column=6, value=user.score)
        ws.cell(row=row_idx, column=7, value=1 if user.test_solved else 0)
        ws.cell(row=row_idx, column=8, value=user.referred_by or "")
        ws.cell(
            row=row_idx,
            column=9,
            value=user.telegram_id,
        )
        ws.cell(
            row=row_idx,
            column=10,
            value=user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else "",
        )

    col_widths = [16, 30, 24, 18, 12, 8, 8, 18, 18, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_referred_users_to_excel(owner, users: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Taklif qilinganlar"

    ws.append(["Taklif qiluvchi", owner.fio or "", owner.telegram_id])
    ws.append([])
    headers = [
        "№",
        "Telegram ID",
        "FIO",
        "Username",
        "Telefon",
        "Ball",
        "Ro'yxatdan o'tganmi",
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=header)

    for row_idx, user in enumerate(users, 4):
        ws.cell(row=row_idx, column=1, value=row_idx - 3)
        ws.cell(row=row_idx, column=2, value=user.telegram_id)
        ws.cell(row=row_idx, column=3, value=user.fio or "")
        ws.cell(row=row_idx, column=4, value=f"@{user.username}" if user.username else "")
        ws.cell(row=row_idx, column=5, value=user.mobile_number or "")
        ws.cell(row=row_idx, column=6, value=user.score)
        ws.cell(row=row_idx, column=7, value="Ha" if user.is_registered else "Yo'q")

    col_widths = [5, 16, 30, 22, 18, 10, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_answers_to_excel(user, session, answers: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Javoblar"

    ws.append(["Foydalanuvchi", user.fio or "", user.telegram_id])
    ws.append(
        [
            "Natija",
            session.score,
            f"/{session.total_questions}",
            "Yakunlangan",
            session.completed_at.strftime("%Y-%m-%d %H:%M") if session.completed_at else "",
        ]
    )
    ws.append([])

    headers = [
        "№",
        "Savol",
        "Tanlangan javob",
        "To'g'ri javob",
        "Natija",
        "Timeout",
        "Vaqt (s)",
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=header)

    for row_idx, answer in enumerate(answers, 5):
        ws.cell(row=row_idx, column=1, value=answer.question_index + 1)
        ws.cell(row=row_idx, column=2, value=answer.question_text or "")
        ws.cell(row=row_idx, column=3, value=answer.selected_answer or "")
        ws.cell(row=row_idx, column=4, value=answer.correct_answer or "")
        ws.cell(row=row_idx, column=5, value="To'g'ri" if answer.is_correct else "Noto'g'ri")
        ws.cell(row=row_idx, column=6, value="Ha" if answer.is_timeout else "Yo'q")
        ws.cell(row=row_idx, column=7, value=answer.time_taken_seconds)

    col_widths = [5, 50, 28, 28, 12, 10, 10]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_test_results_summary_to_excel(rows: list[dict[str, Any]]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test statistikasi"

    headers = [
        "№",
        "Telegram ID",
        "FIO",
        "Username",
        "Session ID",
        "Ball",
        "Jami savol",
        "To'g'ri",
        "Noto'g'ri",
        "Timeout",
        "Jami vaqt (s)",
        "Jami vaqt",
        "Yakunlangan vaqti",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, 2):
        total_time_seconds = int(row.get("total_time_seconds") or 0)
        hours = total_time_seconds // 3600
        minutes = (total_time_seconds % 3600) // 60
        seconds = total_time_seconds % 60
        time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"

        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=row.get("telegram_id"))
        ws.cell(row=row_idx, column=3, value=row.get("fio") or "")
        ws.cell(
            row=row_idx,
            column=4,
            value=f"@{row['username']}" if row.get("username") else "",
        )
        ws.cell(row=row_idx, column=5, value=row.get("session_id"))
        ws.cell(row=row_idx, column=6, value=row.get("score") or 0)
        ws.cell(row=row_idx, column=7, value=row.get("total_questions") or 0)
        ws.cell(row=row_idx, column=8, value=row.get("correct_count") or 0)
        ws.cell(row=row_idx, column=9, value=row.get("incorrect_count") or 0)
        ws.cell(row=row_idx, column=10, value=row.get("timeout_count") or 0)
        ws.cell(row=row_idx, column=11, value=total_time_seconds)
        ws.cell(row=row_idx, column=12, value=time_str)
        completed_at = row.get("completed_at")
        ws.cell(
            row=row_idx,
            column=13,
            value=completed_at.strftime("%Y-%m-%d %H:%M:%S") if completed_at else "",
        )

    col_widths = [5, 16, 30, 20, 12, 8, 10, 8, 10, 10, 14, 12, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_questions_template() -> tuple[io.BytesIO, str]:
    """Generate template file for questions import (.xlsx with example row)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savollar Namuna"

    headers = ["Savol", "To'g'ri javob", "Noto'g'ri 1", "Noto'g'ri 2", "Noto'g'ri 3"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Example row
    ws.append(["Misol savol?", "To'g'ri javob", "Noto'g'ri 1", "Noto'g'ri 2", "Noto'g'ri 3"])

    col_widths = [50, 30, 30, 30, 30]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, "xlsx"


def export_questions_to_excel(questions: list) -> io.BytesIO:
    """Export questions to Excel with columns: Question, Correct, Wrong1, Wrong2, Wrong3"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savollar"

    headers = ["Savol", "To'g'ri javob", "Noto'g'ri 1", "Noto'g'ri 2", "Noto'g'ri 3"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, q in enumerate(questions, 2):
        ws.cell(row=row_idx, column=1, value=q.text or "")
        ws.cell(row=row_idx, column=2, value=q.correct_answer or "")
        ws.cell(row=row_idx, column=3, value=q.answer_2 or "")
        ws.cell(row=row_idx, column=4, value=q.answer_3 or "")
        ws.cell(row=row_idx, column=5, value=q.answer_4 or "")

    col_widths = [50, 30, 30, 30, 30]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_questions_from_excel(path: str) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Excel/CSV format:
    Column A: Question text
    Column B: Correct answer
    Column C: Wrong answer 1
    Column D: Wrong answer 2
    Column E: Wrong answer 3

    Returns: (questions_list, errors_list)
    """
    questions = []
    errors = []
    rows = []

    try:
        if path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(list(row) if row else [])
        elif path.endswith('.csv'):
            with open(path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader, None)  # Skip header
                for row in reader:
                    rows.append(row)
        else:
            errors.append("Fayl formati noto'g'ri. Faqat .xlsx yoki .csv qabul qilinadi.")
            return questions, errors
    except Exception as e:
        errors.append(f"Faylni o'qishda xatolik: {e}")
        return questions, errors

    for row_idx, row in enumerate(rows, start=2):
        if not any(row):
            continue

        try:
            text = str(row[0]).strip() if len(row) > 0 and row[0] else ""
            correct = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            w1 = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            w2 = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            w3 = str(row[4]).strip() if len(row) > 4 and row[4] else ""

            if not text:
                errors.append(f"Qator {row_idx}: Savol yo'q.")
                continue
            if not correct:
                errors.append(f"Qator {row_idx}: To'g'ri javob yo'q.")
                continue
            if not w1:
                errors.append(f"Qator {row_idx}: Noto'g'ri javob 1 yo'q.")
                continue
            if not w2:
                errors.append(f"Qator {row_idx}: Noto'g'ri javob 2 yo'q.")
                continue
            if not w3:
                errors.append(f"Qator {row_idx}: Noto'g'ri javob 3 yo'q.")
                continue

            questions.append(
                dict(
                    text=text,
                    correct=correct,
                    wrong_1=w1,
                    wrong_2=w2,
                    wrong_3=w3,
                )
            )
        except Exception as e:
            errors.append(f"Qator {row_idx}: {e}")

    return questions, errors


def import_users_from_excel(path: str) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Excel/CSV format for importing users:

    Standard format (columns A-G):
    Column A: Telegram ID
    Column B: FIO
    Column C: Username
    Column D: Mobile number
    Column E: Referrals count
    Column F: Score
    Column G: Referred by (user ID)

    CSV format (columns A-J):
    Column A: ID
    Column B: FIO
    Column C: Username
    Column D: Telefon
    Column E: Referallar
    Column F: Ball
    Column G: Javoblar
    Column H: Kim taklif qildi (ID)
    Column I: Telegram ID raqami
    Column J: Qo'shilgan vaqti

    Returns: (users_list, errors_list)
    """
    users = []
    errors = []
    rows = []
    header_map: dict[str, int] = {}

    try:
        if path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header:
                header_map = _detect_user_import_header_map(list(header))
            start_row = 2 if header_map else 1
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                rows.append(list(row) if row else [])
        elif path.endswith('.csv'):
            with open(path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if header:
                    header_map = _detect_user_import_header_map(header)
                    if not header_map:
                        rows.append(header)
                for row in reader:
                    rows.append(row)
        else:
            errors.append("Fayl formati noto'g'ri. Faqat .xlsx yoki .csv qabul qilinadi.")
            return users, errors
    except Exception as e:
        errors.append(f"Faylni o'qishda xatolik: {e}")
        return users, errors

    for row_idx, row in enumerate(rows, start=2):
        if not any(row):
            continue

        try:
            if header_map:
                tid_raw = _get_row_value(row, header_map, "telegram_id")
            else:
                tid_raw = _get_legacy_import_value(row, field="telegram_id")

            if not tid_raw or str(tid_raw).strip() == "":
                errors.append(f"Qator {row_idx}: Telegram ID yo'q.")
                continue

            telegram_id = _to_int_value(tid_raw)

            if header_map:
                fio = _clean_text(_get_row_value(row, header_map, "fio"))
                username = _normalize_username(_clean_text(_get_row_value(row, header_map, "username")))
                mobile_number = _clean_text(_get_row_value(row, header_map, "mobile_number"))
                referrals_count = _to_int_value(
                    _get_row_value(row, header_map, "referrals_count"),
                    default=0,
                )
                score = _to_int_value(
                    _get_row_value(row, header_map, "score"),
                    default=0,
                )
                referred_by = _to_int_value(
                    _get_row_value(row, header_map, "referred_by"),
                    default=None,
                )
            else:
                fio = _clean_text(_get_legacy_import_value(row, field="fio"))
                username = _normalize_username(
                    _clean_text(_get_legacy_import_value(row, field="username"))
                )
                mobile_number = _clean_text(_get_legacy_import_value(row, field="mobile_number"))
                referrals_count = _to_int_value(
                    _get_legacy_import_value(row, field="referrals_count"),
                    default=0,
                )
                score = _to_int_value(
                    _get_legacy_import_value(row, field="score"),
                    default=0,
                )
                referred_by = _to_int_value(
                    _get_legacy_import_value(row, field="referred_by"),
                    default=None,
                )

            users.append({
                "telegram_id": telegram_id,
                "fio": fio,
                "username": username,
                "mobile_number": mobile_number,
                "referrals_count": referrals_count,
                "score": score,
                "referred_by": referred_by,
            })
        except ValueError as e:
            errors.append(f"Qator {row_idx}: {e}")
        except Exception as e:
            errors.append(f"Qator {row_idx}: {e}")

    return users, errors


def _normalize_header_name(value: Any) -> str:
    return str(value or "").strip().lower().replace("’", "'")


def _detect_user_import_header_map(header: list[Any]) -> dict[str, int]:
    aliases = {
        "telegram_id": {"telegram id", "telegram id raqami", "tg id", "id telegram"},
        "fio": {"fio", "f.i.o", "full name", "ism familiya"},
        "username": {"username", "user name", "login"},
        "mobile_number": {"telefon", "phone", "phone number", "mobile number", "mobile"},
        "referrals_count": {"referallar", "referals", "referrals", "referrals count"},
        "score": {"ball", "score"},
        "referred_by": {"kim taklif qildi", "kim taklif qildi (id)", "referred by", "referrer id"},
    }
    detected: dict[str, int] = {}
    for index, column_name in enumerate(header):
        normalized = _normalize_header_name(column_name)
        for field_name, names in aliases.items():
            if normalized in names:
                detected[field_name] = index
                break
    return detected


def _get_row_value(row: list[Any], header_map: dict[str, int], field: str) -> Any:
    index = header_map.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def _get_legacy_import_value(row: list[Any], field: str) -> Any:
    csv_like_indexes = {
        "telegram_id": 8,
        "fio": 1,
        "username": 2,
        "mobile_number": 3,
        "referrals_count": 4,
        "score": 5,
        "referred_by": 7,
    }
    standard_indexes = {
        "telegram_id": 0,
        "fio": 1,
        "username": 2,
        "mobile_number": 3,
        "referrals_count": 4,
        "score": 5,
        "referred_by": 6,
    }
    preferred_indexes = csv_like_indexes if len(row) > 8 else standard_indexes
    index = preferred_indexes[field]
    if index >= len(row):
        return None
    return row[index]


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_username(value: str | None) -> str | None:
    if not value:
        return None
    return value.lstrip("@").strip() or None


def _to_int_value(value: Any, default: int | None = None) -> int | None:
    if value is None:
        return default
    text = str(value).strip()
    if not text:
        return default
    return int(float(text))
