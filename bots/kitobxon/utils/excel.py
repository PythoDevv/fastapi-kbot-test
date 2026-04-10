import csv
import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def export_users_to_excel(users: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"

    headers = [
        "№", "Telegram ID", "FIO", "Username",
        "Telefon", "Ball", "Referallar",
        "Test yechganmi", "Ro'yxatdan o'tgan sana",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, user in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=user.telegram_id)
        ws.cell(row=row_idx, column=3, value=user.fio or "")
        ws.cell(row=row_idx, column=4, value=f"@{user.username}" if user.username else "")
        ws.cell(row=row_idx, column=5, value=user.mobile_number or "")
        ws.cell(row=row_idx, column=6, value=user.score)
        ws.cell(row=row_idx, column=7, value=user.referrals_count)
        ws.cell(row=row_idx, column=8, value="Ha" if user.test_solved else "Yo'q")
        ws.cell(
            row=row_idx,
            column=9,
            value=user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else "",
        )

    col_widths = [5, 14, 30, 20, 16, 8, 10, 12, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_questions_template() -> tuple[io.BytesIO, str]:
    """Generate template file for questions import (.xlsx with example row)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savollar Namuna"

    headers = ["Savol", "To'g'ri javob", "Noto'g'ri 1", "Noto'g'ri 2", "Noto'g'ri 3"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Example row
    ws.append(["Misol savol?", "To'g'ri javob", "Noto'g'ri 1", "Noto'g'ri 2", "Noto'g'ri 3"])

    col_widths = [50, 30, 30, 30, 30]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, "xlsx"


def export_questions_to_excel(questions: list) -> io.BytesIO:
    """Export questions to Excel with columns: Question, Correct, Wrong1, Wrong2, Wrong3"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savollar"

    headers = ["Savol", "To'g'ri javob", "Noto'g'ri 1", "Noto'g'ri 2", "Noto'g'ri 3"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, q in enumerate(questions, 2):
        ws.cell(row=row_idx, column=1, value=q.text or "")
        ws.cell(row=row_idx, column=2, value=q.correct_answer or "")
        ws.cell(row=row_idx, column=3, value=q.answer_2 or "")
        ws.cell(row=row_idx, column=4, value=q.answer_3 or "")
        ws.cell(row=row_idx, column=5, value=q.answer_4 or "")

    col_widths = [50, 30, 30, 30, 30]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_questions_from_excel(path: str) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Excel/CSV format:
    Column A: Question text
    Column B: Correct answer
    Column C: Wrong answer 1
    Column D: Wrong answer 2
    Column E: Wrong answer 3

    Returns: (questions_list, errors_list)
    """
    questions = []
    errors = []
    rows = []

    try:
        if path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(list(row) if row else [])
        elif path.endswith('.csv'):
            with open(path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader, None)  # Skip header
                for row in reader:
                    rows.append(row)
        else:
            errors.append("Fayl formati noto'g'ri. Faqat .xlsx yoki .csv qabul qilinadi.")
            return questions, errors
    except Exception as e:
        errors.append(f"Faylni o'qishda xatolik: {e}")
        return questions, errors

    for row_idx, row in enumerate(rows, start=2):
        if not any(row):
            continue

        try:
            text = str(row[0]).strip() if len(row) > 0 and row[0] else ""
            correct = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            w1 = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            w2 = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            w3 = str(row[4]).strip() if len(row) > 4 and row[4] else ""

            if not text:
                errors.append(f"Qator {row_idx}: Savol yo'q.")
                continue
            if not correct:
                errors.append(f"Qator {row_idx}: To'g'ri javob yo'q.")
                continue
            if not w1:
                errors.append(f"Qator {row_idx}: Noto'g'ri javob 1 yo'q.")
                continue
            if not w2:
                errors.append(f"Qator {row_idx}: Noto'g'ri javob 2 yo'q.")
                continue
            if not w3:
                errors.append(f"Qator {row_idx}: Noto'g'ri javob 3 yo'q.")
                continue

            questions.append(
                dict(
                    text=text,
                    correct=correct,
                    wrong_1=w1,
                    wrong_2=w2,
                    wrong_3=w3,
                )
            )
        except Exception as e:
            errors.append(f"Qator {row_idx}: {e}")

    return questions, errors


def import_users_from_excel(path: str) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Excel/CSV format for importing users:

    Standard format (columns A-G):
    Column A: Telegram ID
    Column B: FIO
    Column C: Username
    Column D: Mobile number
    Column E: Referrals count
    Column F: Score
    Column G: Referred by (user ID)

    CSV format (columns A-J):
    Column A: ID
    Column B: FIO
    Column C: Username
    Column D: Telefon
    Column E: Referallar
    Column F: Ball
    Column G: Javoblar
    Column H: Kim taklif qildi (ID)
    Column I: Telegram ID raqami
    Column J: Qo'shilgan vaqti

    Returns: (users_list, errors_list)
    """
    users = []
    errors = []
    rows = []
    is_csv_format = False

    try:
        if path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(list(row) if row else [])
        elif path.endswith('.csv'):
            with open(path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader, None)  # Read header
                # Detect format based on header
                if header and len(header) > 8:
                    header_lower = [h.lower() if h else "" for h in header]
                    # Check if it's CSV format by looking for specific columns
                    if any("telegram" in h for h in header_lower):
                        is_csv_format = True
                for row in reader:
                    rows.append(row)
        else:
            errors.append("Fayl formati noto'g'ri. Faqat .xlsx yoki .csv qabul qilinadi.")
            return users, errors
    except Exception as e:
        errors.append(f"Faylni o'qishda xatolik: {e}")
        return users, errors

    for row_idx, row in enumerate(rows, start=2):
        if not any(row):
            continue

        try:
            if is_csv_format:
                # CSV format: Telegram ID is in column I (index 8)
                tid_raw = row[8] if len(row) > 8 else row[0]  # Fallback to column A
            else:
                # Standard format: Telegram ID is in column A (index 0)
                tid_raw = row[0] if len(row) > 0 else None

            if not tid_raw or str(tid_raw).strip() == "":
                errors.append(f"Qator {row_idx}: Telegram ID yo'q.")
                continue

            telegram_id = int(float(str(tid_raw).strip()))

            if is_csv_format:
                # CSV format mapping
                fio = str(row[1]).strip() if len(row) > 1 and row[1] else None
                username = str(row[2]).strip() if len(row) > 2 and row[2] else None
                mobile_number = str(row[3]).strip() if len(row) > 3 and row[3] else None
                referrals_count = int(float(str(row[4]).strip())) if len(row) > 4 and row[4] else 0
                score = int(float(str(row[5]).strip())) if len(row) > 5 and row[5] else 0
                referred_by = int(float(str(row[7]).strip())) if len(row) > 7 and row[7] else None
            else:
                # Standard format mapping
                fio = str(row[1]).strip() if len(row) > 1 and row[1] else None
                username = str(row[2]).strip() if len(row) > 2 and row[2] else None
                mobile_number = str(row[3]).strip() if len(row) > 3 and row[3] else None
                referrals_count = int(float(str(row[4]).strip())) if len(row) > 4 and row[4] else 0
                score = int(float(str(row[5]).strip())) if len(row) > 5 and row[5] else 0
                referred_by = int(float(str(row[6]).strip())) if len(row) > 6 and row[6] else None

            users.append({
                "telegram_id": telegram_id,
                "fio": fio,
                "username": username,
                "mobile_number": mobile_number,
                "referrals_count": referrals_count,
                "score": score,
                "referred_by": referred_by,
            })
        except ValueError as e:
            errors.append(f"Qator {row_idx}: {e}")
        except Exception as e:
            errors.append(f"Qator {row_idx}: {e}")

    return users, errors
