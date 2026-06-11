import unittest

import openpyxl

from bots.kitobxon.utils.excel import export_top_answers_to_excel


class TopAnswersExportTests(unittest.TestCase):
    def test_export_top_answers_to_excel_writes_flat_rows(self) -> None:
        rows = [
            {
                "rank": 1,
                "telegram_id": 111,
                "fio": "Ali Valiyev",
                "username": "ali",
                "mobile_number": "+998901234567",
                "session_id": 10,
                "score": 8,
                "total_questions": 10,
                "correct_count": 8,
                "incorrect_count": 1,
                "timeout_count": 1,
                "total_time_seconds": 120,
                "completed_at": "2026-05-07 10:00:00",
                "question_number": 1,
                "question_text": "Savol 1",
                "selected_answer": "Javob A",
                "correct_answer": "Javob A",
                "result": "To'g'ri",
                "timeout": "Yo'q",
                "question_time_seconds": 12,
            },
            {
                "rank": 1,
                "telegram_id": 111,
                "fio": "Ali Valiyev",
                "username": "ali",
                "mobile_number": "+998901234567",
                "session_id": 10,
                "score": 8,
                "total_questions": 10,
                "correct_count": 8,
                "incorrect_count": 1,
                "timeout_count": 1,
                "total_time_seconds": 120,
                "completed_at": "2026-05-07 10:00:00",
                "question_number": 2,
                "question_text": "Savol 2",
                "selected_answer": "Javob B",
                "correct_answer": "Javob C",
                "result": "Noto'g'ri",
                "timeout": "Yo'q",
                "question_time_seconds": 15,
            },
        ]

        workbook_bytes = export_top_answers_to_excel(rows)
        workbook = openpyxl.load_workbook(workbook_bytes)
        sheet = workbook.active

        self.assertEqual(sheet.title, "Top 30 javoblar")
        self.assertEqual(sheet.cell(row=1, column=1).value, "Rank")
        self.assertEqual(sheet.cell(row=1, column=5).value, "Telefon")
        self.assertEqual(sheet.cell(row=1, column=14).value, "Savol №")
        self.assertEqual(sheet.cell(row=2, column=2).value, 111)
        self.assertEqual(sheet.cell(row=2, column=5).value, "+998901234567")
        self.assertEqual(sheet.cell(row=2, column=15).value, "Savol 1")
        self.assertEqual(sheet.cell(row=3, column=17).value, "Javob C")
        self.assertEqual(sheet.cell(row=3, column=18).value, "Noto'g'ri")

    def test_export_with_custom_sheet_title(self) -> None:
        workbook_bytes = export_top_answers_to_excel([], sheet_title="Barcha javoblar")
        workbook = openpyxl.load_workbook(workbook_bytes)
        self.assertEqual(workbook.active.title, "Barcha javoblar")


if __name__ == "__main__":
    unittest.main()
